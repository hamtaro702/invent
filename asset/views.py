from datetime import datetime, date
import pytz
datetime.utcnow().replace(tzinfo=pytz.utc)

from django.core.paginator import EmptyPage, Paginator, InvalidPage
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.db.models import Q


from django.views.decorators.cache import cache_page
import openpyxl
import xlwt
from .models import TheViewAsset, AssetInformation, AssetSystem, AssetModel, Category, Branch, AssetBrand, \
    AssetCategory, AssetStatus


# @cache_page(60 * 5)
# Create your views here.

def export_users_xls(request):
    today_date = date.today()
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Inventory_' + str(today_date) + '.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('sheet1')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True


    columns = ['asset_id', 'Status', 'System', 'Category',
               'Host Name', 'User', 'Brand Name', 'Model Name', 'Serial Number',
               'Mac Address Lan', 'Mac Address WiFi', 'IP WiFi', 'IP Lan', 'Asset No.', 'หนังสือสัญญา',
               'Cost Center', 'Department', 'Branch Name', 'Building', 'Floor', 'Project', 'Device Name',
               'IMEI', 'Phone Number', 'SIM', 'PO', 'Invoice', 'Stock In', 'Stock Out', 'Remark', 'Create Date',
               'Create By', 'Modified Date', 'Modified By', 'สถานะเครื่อง', 'windows_edition', 'win_product_key',
               'office_edition', 'office_product_key', 'Line', 'VPN', 'CPU', 'RAM', 'Storage', 'Filepath']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = TheViewAsset.objects.all().values_list('asset_id', 'status_name', 'system_name', 'category_name',
                                                  'host_name',
                                                  'user', 'brand_name', 'model_name', 'serial_number',
                                                  'mac_address_lan', 'mac_address_wifi',
                                                  'ip_wifi', 'ip_lan', 'asset', 'bookno', 'costcenter', 'location',
                                                  'branch_name', 'building',
                                                  'floor', 'project', 'device_name', 'imei', 'phone', 'sim', 'po',
                                                  'invoice', 'date_stockin',
                                                  'date_stockout', 'remark', 'createdate', 'createby', 'modifieddate',
                                                  'modifiedby', 'status_name',
                                                  'windows_edition', 'win_product_key', 'office_edition',
                                                  'office_product_key', 'line',
                                                  'vpn', 'cpu', 'ram', 'storage', 'filepath'
                                                  ).order_by('-asset_id')
    rows = [[x.strftime("%Y-%m-%d %H:%M") if isinstance(x, datetime) else x for x in row] for row in rows]
    rows = [[x.strftime("%Y-%m-%d") if isinstance(x, date) else x for x in row] for row in rows]
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)


    wb.save(response)
    return response


def readExcel(request):
    if "GET" == request.method:
        return render(request, 'excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        assetlist = list()
        statuslist = list()
        systemlist = list()
        categorylist = list()
        hostnamelist = list()
        userlist = list()
        brandlist = list()
        modellist = list()
        serialnumberlist = list()
        maclanlist = list()
        macwifilist = list()
        ipwifilist = list()
        iplanlist = list()
        assetnolist = list()
        booknolist = list()
        costcenterlist = list()
        departmentlist = list()
        branchlist = list()
        buildinglist = list()
        floorlist = list()
        projectlist = list()
        devicelist = list()
        imeilist = list()
        phonenumberlist = list()
        simlist = list()
        polist = list()
        invoicelist = list()
        stockinlist = list()
        stockoutlist = list()
        remarklist = list()
        createdatelist = list()
        createbylist = list()
        modifieddatelist = list()
        modifiedbylist = list()
        conditionnolist = list()
        windows_editionlist = list()
        win_product_keylist = list()
        office_editionlist = list()
        office_product_keylist = list()
        linelist = list()
        vpnlist = list()
        cpulist = list()
        ramlist = list()
        storagelist = list()
        filelist = list()

        # iterating over the rows and
        # getting value from each cell in row
        i = 0
        j = 0
        for row in worksheet.iter_rows():
            row_data = list()
            if i != 0:
                print("i:" + str(i))
                for cell in row:
                    print("j:" + str(j))
                    row_data.append(str(cell.value))
                    if j == 0:
                        assetlist.append(str(cell.value))
                    elif j == 1:
                        statuslist.append(str(cell.value))
                    elif j == 2:
                        systemlist.append(str(cell.value))
                    elif j == 3:
                        categorylist.append(str(cell.value))
                    elif j == 4:
                        hostnamelist.append(str(cell.value))
                    elif j == 5:
                        userlist.append(str(cell.value))
                    elif j == 6:
                        brandlist.append(str(cell.value))
                    elif j == 7:
                        modellist.append(str(cell.value))
                    elif j == 8:
                        serialnumberlist.append(str(cell.value))
                    elif j == 9:
                        maclanlist.append(str(cell.value))
                    elif j == 10:
                        macwifilist.append(str(cell.value))
                    elif j == 11:
                        ipwifilist.append(str(cell.value))
                    elif j == 12:
                        iplanlist.append(str(cell.value))
                    elif j == 13:
                        assetnolist.append(str(cell.value))
                    elif j == 14:
                        booknolist.append(str(cell.value))
                    elif j == 15:
                        costcenterlist.append(str(cell.value))
                    elif j == 16:
                        departmentlist.append(str(cell.value))
                    elif j == 17:
                        branchlist.append(str(cell.value))
                    elif j == 18:
                        buildinglist.append(str(cell.value))
                    elif j == 19:
                        floorlist.append(str(cell.value))
                    elif j == 20:
                        projectlist.append(str(cell.value))
                    elif j == 21:
                        devicelist.append(str(cell.value))
                    elif j == 22:
                        imeilist.append(str(cell.value))
                    elif j == 23:
                        phonenumberlist.append(str(cell.value))
                    elif j == 24:
                        simlist.append(str(cell.value))
                    elif j == 25:
                        polist.append(str(cell.value))
                    elif j == 26:
                        invoicelist.append(str(cell.value))
                    elif j == 27:
                        stockinlist.append(str(cell.value))
                    elif j == 28:
                        stockoutlist.append(str(cell.value))
                    elif j == 29:
                        remarklist.append(str(cell.value))
                    elif j == 30:
                        createdatelist.append(str(cell.value))
                    elif j == 31:
                        createbylist.append(str(cell.value))
                    elif j == 32:
                        modifieddatelist.append(str(cell.value))
                    elif j == 33:
                        modifiedbylist.append(str(cell.value))
                    elif j == 34:
                        conditionnolist.append(str(cell.value))
                    elif j == 35:
                        windows_editionlist.append(str(cell.value))
                    elif j == 36:
                        win_product_keylist.append(str(cell.value))
                    elif j == 37:
                        office_editionlist.append(str(cell.value))
                    elif j == 38:
                        office_product_keylist.append(str(cell.value))
                    elif j == 39:
                        linelist.append(str(cell.value))
                    elif j == 40:
                        vpnlist.append(str(cell.value))
                    elif j == 41:
                        cpulist.append(str(cell.value))
                    elif j == 42:
                        ramlist.append(str(cell.value))
                    elif j == 43:
                        storagelist.append(str(cell.value))
                    elif j == 44:
                        filelist.append(str(cell.value))
                    j = j + 1
                excel_data.append(row_data)
                j = 0
                data = AssetInformation.objects.get(asset_id=assetlist[i - 1])
                print(assetlist[i - 1])
                if Branch.objects.filter(branch_name=branchlist[i - 1]):
                    branch = Branch.objects.get(branch_name__exact=branchlist[i - 1])
                    data.branch_id = branch.branch_id
                if AssetStatus.objects.filter(status_name=statuslist[i - 1]):
                    status = AssetStatus.objects.get(status_name__exact=statuslist[i - 1])
                    data.status_id = status.status_id
                if AssetSystem.objects.filter(system_name=systemlist[i - 1]):
                    system = AssetSystem.objects.get(system_name__exact=systemlist[i - 1])
                    data.system_id = system.system_id
                if AssetCategory.objects.filter(category_name=categorylist[i - 1]):
                    print(categorylist[i - 1])
                    category = AssetCategory.objects.get(category_name__exact=categorylist[i - 1])
                    data.category_id = category.category_id

                data.host_name = hostnamelist[i - 1]
                data.user = userlist[i - 1]

                if AssetBrand.objects.filter(brand_name=brandlist[i - 1]):
                    brand = AssetBrand.objects.get(brand_name=brandlist[i - 1])
                    data.brand_id = brand.brand_id
                if AssetModel.objects.filter(model_name=modellist[i - 1]):
                    model = AssetModel.objects.get(model_name=modellist[i - 1])
                    data.model_id = model.model_id

                data.serial_number = serialnumberlist[i - 1]
                data.mac_address_lan = maclanlist[i - 1]
                data.mac_address_wifi = macwifilist[i - 1]
                data.ip_wifi = ipwifilist[i - 1]
                data.ip_lan = iplanlist[i - 1]
                data.asset = assetnolist[i - 1]
                data.bookno = booknolist[i - 1]
                data.costcenter = costcenterlist[i - 1]
                data.location = departmentlist[i - 1]
                data.building = buildinglist[i - 1]
                data.floor = floorlist[i - 1]
                data.project = projectlist[i - 1]
                data.device_name = devicelist[i - 1]
                data.imei = imeilist[i - 1]
                data.phone = phonenumberlist[i - 1]
                data.sim = simlist[i - 1]
                data.po = polist[i - 1]
                data.invoice = invoicelist[i - 1]
                data.date_stockin = stockinlist[i - 1]
                data.date_stockout = stockoutlist[i - 1]
                data.remark = remarklist[i - 1]
                data.createdate = createdatelist[i - 1]
                data.createby = createbylist[i - 1]
                data.modifieddate = datetime.now()
                data.modifiedby = modifiedbylist[i - 1]
                data.condition_no = conditionnolist[i - 1]
                data.windows_edition = windows_editionlist[i - 1]
                data.win_product_key = win_product_keylist[i - 1]
                data.office_edition = office_editionlist[i - 1]
                data.office_product_key = office_product_keylist[i - 1]
                data.line = linelist[i - 1]
                data.vpn = vpnlist[i - 1]
                data.cpu = cpulist[i - 1]
                data.ram = ramlist[i - 1]
                data.storage = storagelist[i - 1]
                data.filepath = filelist[i - 1]

                data.save()
            i = i + 1

        return render(request, 'upload.html', {"excel_data": excel_data})


def index(request):
    # user = request.POST['user']

    return render(request, "index.html")


def search(request):
    assetcode = request.POST['assetcode']
    user = request.POST['user']

    print("assetcode:" + assetcode)
    if assetcode != '':
        data = TheViewAsset.objects.all().filter(asset__contains=assetcode).order_by('-asset_id')
        return render(request, "index.html", {'data': data})
    elif user != '':
        data = TheViewAsset.objects.all().filter(user__contains=user).order_by('-asset_id')
        return render(request, "index.html", {'data': data})
    else:
        data = TheViewAsset.objects.all().order_by('-asset_id')
        return render(request, "index.html", {'data': data})


def searchGlobal(request):
    query = request.GET.get('q')
    if query:
        paginator = Paginator(TheViewAsset.objects.filter(
            Q(user__contains=query) | Q(asset__contains=query) | Q(host_name__contains=query) | Q(
                brand_name__contains=query) | Q(model_name__contains=query) | Q(category_name__contains=query) | Q(
                branch_name__contains=query) | Q(status_name__contains=query) | Q(serial_number__contains=query)
            | Q(bookno__contains=query) | Q(remark__contains=query)
        ).distinct().order_by('-asset_id'), 5)

        try:
            page = int(request.GET.get('page', '1'))
        except:
            page = 1

        try:
            blogs = paginator.page(page)
        except(EmptyPage, InvalidPage):
            blogs = paginator.page(1)

            # Get the index of the current page
        index = blogs.number - 1  # edited to something easier without index
        # This value is maximum index of your pages, so the last page - 1
        max_index = len(paginator.page_range)
        # You want a range of 7, so lets calculate where to slice the list
        start_index = index - 5 if index >= 5 else 0
        end_index = index + 5 if index <= max_index - 5 else max_index
        # Get our new page range. In the latest versions of Django page_range returns
        # an iterator. Thus pass it to list, to make our slice possible again.
        page_range = list(paginator.page_range)[start_index:end_index]

        return render(request, 'index.html', {
            'data': blogs,
            'page_range': page_range,
        })
    else:
        return render(request, 'index.html')


def fetchALL(request):
    paginator = Paginator(TheViewAsset.objects.all().order_by('-asset_id'), 5)
    try:
        page = int(request.GET.get('page', '1'))
    except:
        page = 1

    try:
        blogs = paginator.page(page)
    except(EmptyPage, InvalidPage):
        blogs = paginator.page(1)

        # Get the index of the current page
    index = blogs.number - 1  # edited to something easier without index
    # This value is maximum index of your pages, so the last page - 1
    max_index = len(paginator.page_range)
    # You want a range of 7, so lets calculate where to slice the list
    start_index = index - 7 if index >= 7 else 0
    end_index = index + 7 if index <= max_index - 7 else max_index
    # Get our new page range. In the latest versions of Django page_range returns
    # an iterator. Thus pass it to list, to make our slice possible again.
    page_range = list(paginator.page_range)[start_index:end_index]

    return render(request, 'index.html', {
        'data': blogs,
        'page_range': page_range,
    })


def assetDelete(request):
    url = "fetchALL"
    query = request.GET.get('q')
    assetid = request.GET.get('assetid')
    AssetInformation.objects.filter(asset_id=assetid).delete()

    if query == "None" or query is None:
        url = "fetchALL"
        return HttpResponseRedirect(url)
    else:
        url = "searchGlobal?q=" + query
        return HttpResponseRedirect(url)


def assetAdd(request):
    systems = AssetSystem.objects.all().order_by('system_name')
    models = AssetModel.objects.all().order_by('model_name')
    categorys = AssetCategory.objects.all().order_by('category_name')
    brands = AssetBrand.objects.all().order_by('brand_name')
    branchs = Branch.objects.all().order_by('branch_name')
    status = AssetStatus.objects.all()
    return render(request, "assetAdd.html",
                  {'systems': systems, 'models': models, 'categorys': categorys, 'brands': brands, 'branchs': branchs,
                   'status': status})


def asset(request):
    if request.GET.get('assetid'):
        assetid = request.GET.get('assetid')
        query = request.GET.get('q')
        print(query)
        print(assetid)
        systems = AssetSystem.objects.all().order_by('system_name')
        models = AssetModel.objects.all().order_by('model_name')
        categorys = AssetCategory.objects.all().order_by('category_name')
        brands = AssetBrand.objects.all().order_by('brand_name')
        branchs = Branch.objects.all().order_by('branch_name')
        status = AssetStatus.objects.all()
        if assetid != '':
            data = AssetInformation.objects.all().filter(asset_id=assetid)
            return render(request, "asset.html",
                          {'q': query, 'data': data, 'systems': systems, 'models': models, 'categorys': categorys,
                           'brands': brands, 'branchs': branchs, 'status': status})


def saveFrom(request):
    asset_id = request.POST['asset_id']
    query = request.GET.get('q')
    print(request.POST['status_id'])
    if asset_id:
        data = AssetInformation.objects.get(asset_id=asset_id)
        data.system_id = request.POST['system']
        data.remark = request.POST['remark']
        data.category_id = request.POST['category']
        data.host_name = request.POST['hostname']
        data.project = request.POST['project']
        data.device_name = request.POST['devicename']
        data.user = request.POST['user']
        data.brand_id = request.POST['brand']
        data.model_id = request.POST['model']
        data.serial_number = request.POST['serial_number']
        data.imei = request.POST['imei']
        data.phone = request.POST['phone']
        data.sim = request.POST['sim']
        data.mac_address_lan = request.POST['mac_address_lan']
        data.mac_address_wifi = request.POST['mac_address_wifi']
        data.ip_wifi = request.POST['ip_wifi']
        data.asset = request.POST['asset']
        data.location = request.POST['location']
        data.branch_id = request.POST['branch']
        data.building = request.POST['building']
        data.floor = request.POST['floor']
        data.status_id = request.POST['status_id']
        data.po = request.POST['po']
        data.invoice = request.POST['invoice']
        data.bookno = request.POST['Bookno']
        data.condition_no = request.POST['condition_no']
        data.windows_edition = request.POST['windows_edition']
        data.win_product_key = request.POST['win_product_key']
        data.office_edition = request.POST['office_edition']
        data.line = request.POST['line']
        data.office_product_key = request.POST['office_product_key']
        data.vpn = request.POST['vpn']
        data.cpu = request.POST['cpu']
        data.ram = request.POST['ram']
        data.storage = request.POST['storage']
        data.filepath = request.POST['filepath']
        data.costcenter = request.POST['costcenter']

        data.modifieddate = datetime.now()
        data.save()
        print(query)
        if query == "None" or query is None:
            url = "fetchALL"
            return HttpResponseRedirect(url)
        else:
            url = "searchGlobal?q=" + query
            return HttpResponseRedirect(url)
    else:
        print(request.POST['system'])
        data = AssetInformation(system_id=request.POST['system'], remark=request.POST['remark'],
                                category_id=request.POST['category'], host_name=request.POST['hostname'],
                                project=request.POST['project'], device_name=request.POST['devicename'],
                                user=request.POST['user'], brand_id=request.POST['brand'],
                                model_id=request.POST['model'], serial_number=request.POST['serial_number']
                                , imei=request.POST['imei'], phone=request.POST['phone'], sim=request.POST['sim'],
                                mac_address_lan=request.POST['mac_address_lan']
                                , mac_address_wifi=request.POST['mac_address_wifi'], ip_wifi=request.POST['ip_wifi'],
                                asset=request.POST['asset'],
                                location=request.POST['location'], branch_id=request.POST['branch'],
                                status_id=request.POST['status_id'],
                                building=request.POST['building'], floor=request.POST['floor'],
                                po=request.POST['po'], invoice=request.POST['invoice'], bookno=request.POST['Bookno'],
                                condition_no=request.POST['condition_no'],
                                windows_edition=request.POST['windows_edition'],
                                win_product_key=request.POST['win_product_key'],
                                office_edition=request.POST['office_edition'],
                                modifieddate=datetime.now(),
                                line=request.POST['line'], office_product_key=request.POST['office_product_key'],
                                vpn=request.POST['vpn'], cpu=request.POST['cpu'],
                                ram=request.POST['ram'], storage=request.POST['storage'],
                                filepath=request.POST['filepath'], costcenter=request.POST['costcenter'])
        data.save()

        if query == "None" or query is None:
            url = "fetchALL"
            return HttpResponseRedirect(url)
        else:
            url = "searchGlobal?q=" + query
            return HttpResponseRedirect(url)


def createFrom(request):
    return render(request, "createFrom.html")


def editForm(request):
    return render(request, "editForm.html")


def addForm(request):
    user = request.POST['user']
    password = request.POST['password']
    return render(request, "addForm.html", {'user': user, 'password': password})


def upload(request):
    return render(request, "upload.html")
